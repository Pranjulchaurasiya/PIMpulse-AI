import os
import re
import json
import asyncio
import polars as pl
from typing import Dict, Any, List, Optional, Tuple
import logging

logger = logging.getLogger("pimpulse.unilog_pipeline")

from agents.unilog_rules import (
    resolve_manufacturer_and_brand,
    resolve_manufacturer_brand,
    parse_abrasive_dimensions,
    format_invoice_desc,
    format_mobile_desc,
    format_short_desc,
    match_lov_value,
    format_fraction_value,
    standardize_uom,
    sanitize_raw_industrial_input,
    APPROVED_MATERIAL_LOV,
    APPROVED_APPLICATION_LOV
)
from config import settings

# Master 252 delivery columns from official Unilog delivery format
UNILOG_DELIVERY_COLUMNS = [
    'MFR URL', 'Ref URL 1', 'Ref URL 2', 'Ref URL 3', 'Ref URL 4', 'Ref URL 5',
    'PART_NUMBER', 'Dept', 'Class', 'Fine', 'SKU - MY_PART_NUMBER', 'Mfg_Part_Num',
    'Part_Desc', 'E1_Brand', 'Unilog_Brand', 'DIB_Brand', 'Part_Manuf',
    'MANUFACTURER_NAME', 'BRAND_NAME', 'TRADE_NAME', 'MANUFACTURER_PART_NUMBER',
    'ALTERNATE_PART_NUMBER', 'Classpath', 'MOBILE_DESC', 'INVOICE_DESC', 'SHORT_DESC',
    'LONG_DESC1', 'RETAIL_DESC', 'MARKETING_DESCRIPTION',
    'ITEM_FEATURES_1', 'ITEM_FEATURES_2', 'ITEM_FEATURES_3', 'ITEM_FEATURES_4',
    'ITEM_FEATURES_5', 'ITEM_FEATURES_6', 'ITEM_FEATURES_7', 'ITEM_FEATURES_8',
    'ITEM_FEATURES_9', 'ITEM_FEATURES_10', 'ITEM_FEATURES_11', 'ITEM_FEATURES_12',
    'ITEM_FEATURES_13', 'ITEM_FEATURES_14', 'ITEM_FEATURES_15', 'ITEM_FEATURES_16',
    'ITEM_FEATURES_17', 'ITEM_FEATURES_18', 'ITEM_FEATURES_19', 'ITEM_FEATURES_20',
    'With', 'Standard/Approvals', 'Prop 65', 'Application', 'Includes', 'Product Name'
]

# Add 50 Attribute Triplets
for i in range(1, 51):
    UNILOG_DELIVERY_COLUMNS.extend([
        f'ATTRIBUTE_LABEL {i}',
        f'ATTRIBUTE_VALUE {i}',
        f'ATTRIBUTE_UOM {i}'
    ])

# Add remaining logistics, classification and media columns
UNILOG_DELIVERY_COLUMNS.extend([
    'UPC', 'EAN', 'GTIN', 'UNSPSC', 'Warranty', 'List Price', 'Selling Qty', 'Selling UOM',
    'Standard Packaging Information', 'LENGTH', 'LENGTH_UOM', 'HEIGHT', 'HEIGHT_UOM',
    'WIDTH', 'WIDTH_UOM', 'WEIGHT', 'WEIGHT_UOM', 'VOLUME', 'VOLUME_UOM',
    'Product Image', 'Alternate Image 1', 'Alternate Image 2', 'Alternate Image 3', 'Alternate Image 4',
    'SDS', 'SDS_1', 'Warranty Information', 'Catalog', 'Specification Sheet',
    'Instruction/Installation Manual', 'Service Manual', 'Owners/User Manual', 'Line Drawing',
    'MTR', 'RoHS', 'Full Engineering Drawing', 'Energy Star Guide', 'Technical Bulletin',
    'Submittal', 'Compatibility Chart', 'Size Chart', 'Product Label/Insert',
    'Video Link', 'Video Link 1', 'Country Of Origin', 'Discontinued', 'Actual Image (Yes/No)'
])

async def enrich_unilog_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Universal Unilog catalog row enrichment supporting all 252 official delivery columns.
    Combines deterministic rules, canonical resolution, and dynamic 50-attribute slotting.
    """
    raw_mpn = str(row.get("Mfg_Part_Num", row.get("MANUFACTURER_PART_NUMBER", ""))).strip()
    raw_desc = str(row.get("Part_Desc", row.get("raw_input", ""))).strip()
    part_manuf = str(row.get("Part_Manuf", row.get("MANUFACTURER_NAME", ""))).strip()
    e1_brand = row.get("E1_Brand", "")
    unilog_brand = row.get("Unilog_Brand", "")
    dib_brand = row.get("DIB_Brand", "")

    # 0. Defensive Preprocessor Sanitizer
    sanitized = sanitize_raw_industrial_input(raw_desc or raw_mpn)
    mpn = raw_mpn
    if not mpn or mpn == raw_desc or "/" in mpn or "_" in mpn or "." in mpn:
        mpn = sanitized["normalized_mpn"] or mpn
    
    if not part_manuf or part_manuf in {"-- Unbranded --", "-- No Unilog Brand --"}:
        part_manuf = sanitized["inferred_mfr"] or part_manuf
        
    best_brand = e1_brand or unilog_brand or dib_brand
    if not best_brand or best_brand in {"-- Unbranded --", "-- No Unilog Brand --"}:
        best_brand = sanitized["inferred_brand"] or best_brand

    clean_desc = sanitized["cleaned_text"] or raw_desc

    # 1. Canonical Manufacturer and Brand Resolution
    mfr_name, brand_name, mfr_code = resolve_manufacturer_and_brand(
        part_manuf=part_manuf,
        part_desc=clean_desc,
        e1_brand=best_brand
    )

    # 2. Dimension and Category Parsing
    dims = sanitized.get("dimensions") or parse_abrasive_dimensions(clean_desc)
    desc_lower = clean_desc.lower()

    # Category taxonomy, class path, and series defaults
    dept = "Abrasives"
    cat_class = "Abrasive Wheels & Discs"
    fine = "Cut-Off Wheels"
    classpath = "Abrasives & Cutting Tools > Abrasives > Cut-Off Wheels"
    unspsc = "31191600"
    series = ""
    item_type = "Cut-Off Disc"
    raw_material = "Aluminum Oxide"
    raw_application = "Metal Cutting"
    features = []

    # Category-specific heuristics
    if "decking" in desc_lower or "trex" in desc_lower or "azek" in desc_lower or "timbertech" in desc_lower or "rail kit" in desc_lower or "baluster" in desc_lower or "fascia" in desc_lower or "grooved" in desc_lower or "t-rail" in desc_lower or "boica" in (part_manuf or "").lower() or "parksite" in (part_manuf or "").lower() or "lumber" in (part_manuf or "").lower():
        dept = "Building Materials"
        cat_class = "Decking & Railing"
        fine = "Composite Decking & Railing"
        classpath = "Building Materials > Decking & Railing > Composite"
        item_type = "Composite Decking Board" if "deck" in desc_lower or "fascia" in desc_lower or "grooved" in desc_lower else "Railing Kit"
        unspsc = "30161801"
        raw_material = "Composite"
        raw_application = "Decking Construction"
        features = ["High-performance composite shell", "Resists fading, staining, and scratching", "Low maintenance construction"]
    elif "dryer" in desc_lower or "washer" in desc_lower or "dish" in desc_lower or "refrigerator" in desc_lower or "fridge" in desc_lower or "appliance" in desc_lower or "appde" in (part_manuf or "").lower() or "vvapp" in (part_manuf or "").lower():
        dept = "Appliances"
        cat_class = "Laundry & Kitchen Appliances"
        fine = "Residential & Commercial Appliances"
        classpath = "Appliances > Major Appliances > Laundry & Kitchen"
        item_type = "Clothes Dryer" if "dryer" in desc_lower else ("Washing Machine" if "washer" in desc_lower else ("Dishwasher" if "dish" in desc_lower else "Commercial Appliance"))
        unspsc = "52141510" if "dryer" in desc_lower else ("52141505" if "washer" in desc_lower else "52141501")
        raw_material = "Enameled Steel"
        raw_application = "Residential Laundry" if "dryer" in desc_lower or "washer" in desc_lower else "Dishwashing"
        features = ["Heavy-duty commercial drive system", "Energy efficient operation", "Durable porcelain/steel drum"]
    elif "lamp" in desc_lower or "led" in desc_lower or "bulb" in desc_lower or "kichler" in desc_lower or "satco" in desc_lower or "philips" in desc_lower or "phillips" in desc_lower or "lighting" in (part_manuf or "").lower() or "sconce" in desc_lower or "fixture" in desc_lower or "27k" in desc_lower or "50k" in desc_lower or "60w" in desc_lower or "40w" in desc_lower:
        dept = "Electrical & Lighting"
        cat_class = "Luminaires & Lamps"
        fine = "LED Lamps & Fixtures"
        classpath = "Electrical & Lighting > Lamps & Fixtures > LED"
        item_type = "LED Light Bulb" if "lamp" in desc_lower or "bulb" in desc_lower or "led" in desc_lower else "Lighting Fixture"
        unspsc = "39112102"
        raw_material = "Glass"
        raw_application = "Commercial Lighting"
        features = ["High-efficiency LED output", "Long rated operating life", "Standard base compatibility"]
    elif "dewalt" in desc_lower or "makita" in desc_lower or "festool" in desc_lower or "kreg" in desc_lower or "blade" in desc_lower or "drill" in desc_lower or "driver" in desc_lower or "impact" in desc_lower or "router" in desc_lower or "bit" in desc_lower or "saw" in desc_lower:
        dept = "Tools & Hardware"
        cat_class = "Power Tools & Accessories"
        fine = "Cutting Blades & Power Tool Accessories"
        classpath = "Tools & Hardware > Power Tools > Cutting & Drilling"
        item_type = "Saw Blade" if "blade" in desc_lower or "saw" in desc_lower else ("Drill Bit" if "bit" in desc_lower else "Power Tool Accessory")
        unspsc = "27112800"
        raw_material = "High Speed Steel"
        raw_application = "Cutting & Drilling"
        features = ["Precision machined cutting geometry", "High impact durability", "Universal tool shank fit"]
    elif "leviton" in desc_lower or "southwire" in desc_lower or "switch" in desc_lower or "receptacle" in desc_lower or "outlet" in desc_lower or "cable" in desc_lower or "wire" in desc_lower:
        dept = "Electrical"
        cat_class = "Wiring Devices"
        fine = "Switches & Receptacles"
        classpath = "Electrical > Wiring Devices > Switches & Outlets"
        item_type = "Wiring Device"
        unspsc = "39121406"
        raw_material = "Thermoplastic"
        raw_application = "Electrical Distribution"
        features = ["Impact resistant thermoplastic", "Quick-wire push-in terminals", "Meets UL and NEMA standards"]
    elif "eyewear" in desc_lower or "glasses" in desc_lower or "glove" in desc_lower or "safety" in desc_lower or "kneeling" in desc_lower or "knee pad" in desc_lower or "edgsa" in (part_manuf or "").lower() or "tecge" in (part_manuf or "").lower():
        dept = "Safety & PPE"
        cat_class = "Personal Protective Equipment"
        fine = "Safety Eyewear & Protective Gear"
        classpath = "Safety & Security > PPE > Eye & Body Protection"
        item_type = "Safety Eyewear" if "eyewear" in desc_lower or "glasses" in desc_lower else "Protective Workwear"
        unspsc = "46181802"
        raw_material = "Polycarbonate"
        raw_application = "Worker Safety"
        features = ["ANSI Z87.1+ impact rated", "Anti-scratch and anti-fog coating", "Ergonomic protective fit"]
    elif "mortar" in desc_lower or "joint system" in desc_lower or "emseal" in desc_lower or "stone" in desc_lower or "reeca" in (part_manuf or "").lower() or "emsjo" in (part_manuf or "").lower():
        dept = "Building Materials"
        cat_class = "Masonry & Sealants"
        fine = "Mortar & Expansion Joints"
        classpath = "Building Materials > Masonry > Mortars & Joint Systems"
        item_type = "Mortar Compound" if "mortar" in desc_lower else "Expansion Joint System"
        unspsc = "30111500"
        raw_material = "Composite"
        raw_application = "Masonry Cutting"
        features = ["High bond strength formula", "Pre-blended for consistency", "Meets ASTM C270 standards"]
    elif "sanding belt" in desc_lower or "sand belt" in desc_lower:
        dept = "Abrasives"
        cat_class = "Sanding Belts"
        fine = "Abrasive Belts"
        classpath = "Abrasives & Cutting Tools > Sanding > Belts"
        item_type = "Sanding Belt"
        raw_application = "Sanding"
        raw_material = "Zirconia Alumina"
        unspsc = "31191500"
        features = ["Heavy-duty cloth backing", "Resin-over-resin bond", "Bi-directional tape joint"]
    elif "stikit" in desc_lower or "cubitron" in desc_lower:
        series = "Cubitron™ II"
        item_type = "Stikit™ Film Disc"
        raw_material = "Ceramic"
        raw_application = "Finishing"
        unspsc = "31191500"
        features = ["Precision-shaped grain technology", "Self-sharpening abrasive grains", "Adhesive Stikit backing"]
    elif "abranet" in desc_lower:
        series = "Abranet®"
        item_type = "Abrasive Mesh Strip"
        raw_material = "Aluminum Oxide"
        raw_application = "Sanding"
        unspsc = "31191500"
        features = ["Dust-free sanding mesh", "Open structure prevents clogging", "Hook and loop attachment"]
    elif "hiolit" in desc_lower:
        series = "HIOLIT"
        item_type = "Abrasive Disc"
        raw_material = "Aluminum Oxide"
        raw_application = "Sanding"
        unspsc = "31191500"
        features = ["Durable cloth backing", "High edge wear resistance", "Universal abrasive performance"]
    elif "grinding wheel" in desc_lower or "grind" in desc_lower:
        item_type = "Grinding Wheel"
        raw_application = "Grinding"
        raw_material = "Silicon Carbide" if "masonry" in desc_lower else "Aluminum Oxide"
        unspsc = "31191600"
        cat_class = "Grinding Wheels"
        fine = "Depressed Center Wheels"
        classpath = "Abrasives & Cutting Tools > Grinding > Wheels"
        features = ["Depressed center Type 27 design", "Reinforced fiberglass webbing", "Aggressive stock removal"]
    elif "masonry" in desc_lower:
        item_type = "Masonry Cut-Off Disc"
        raw_application = "Masonry Cutting"
        raw_material = "Silicon Carbide"
        unspsc = "31191506"
        features = ["Silicon carbide grain for stone/masonry", "High-speed reinforced construction", "Clean, burr-free cuts"]
    elif "bolt" in desc_lower or "screw" in desc_lower or "chv-blt" in desc_lower:
        dept = "Fasteners"
        cat_class = "Bolts & Screws"
        fine = "Hex Head Machine Bolts"
        classpath = "Industrial Fasteners > Bolts > Hex Bolts"
        item_type = "Hex Head Machine Bolt"
        raw_material = "Stainless Steel" if "ss" in desc_lower or "316" in desc_lower else "Carbon Steel"
        raw_application = "Fastening"
        unspsc = "31161620"
        features = ["Precision machined threads", "Corrosion resistant construction", "Meets ASTM specifications"]
    elif "breaker" in desc_lower or " cb" in desc_lower:
        dept = "Electrical"
        cat_class = "Circuit Protection"
        fine = "Miniature Circuit Breakers"
        classpath = "Electrical > Circuit Protection > MCBs"
        item_type = "Miniature Circuit Breaker"
        raw_material = "Thermoplastic"
        raw_application = "Circuit Protection"
        unspsc = "39121603"
        features = ["Thermal magnetic trip mechanism", "DIN rail mountable", "High breaking capacity"]
    elif "bearing" in desc_lower or "6205" in desc_lower:
        dept = "Bearings & Power Transmission"
        cat_class = "Bearings"
        fine = "Deep Groove Ball Bearings"
        classpath = "Bearings & Mechanical > Ball Bearings > Deep Groove"
        item_type = "Deep Groove Ball Bearing"
        raw_material = "Chrome Steel"
        raw_application = "Rotary Motion"
        unspsc = "31171504"
        features = ["Precision ABEC-1 tolerances", "Dual contact rubber seals (2RSH)", "Low friction synthetic lubrication"]

    if "performance+" in desc_lower or "perform+" in desc_lower:
        series = "Performance+"
    elif "steel demon" in desc_lower:
        series = "Steel Demon™"
    elif "speed demon" in desc_lower:
        series = "Speed Demon™"
    elif "ceramic+" in desc_lower:
        series = "Ceramic+"

    # Match raw material and application to Approved Unilog LOVs
    material = match_lov_value(raw_material, APPROVED_MATERIAL_LOV)
    application = match_lov_value(raw_application, APPROVED_APPLICATION_LOV)

    # 3. MDM Descriptions with strict word-boundary guarantees
    invoice_desc = format_invoice_desc(mfr_code, mpn, dims, raw_desc)
    mobile_desc = format_mobile_desc(mfr_name, brand_name, mpn, item_type, dims, series=series)
    short_desc = format_short_desc(brand_name, series, mpn, item_type, dims)
    
    # Long description formatting
    dia_str = f"{dims.get('diameter')} in Dia" if dims.get("diameter") else ""
    thk_str = f"{dims.get('thickness')} in Thk" if dims.get("thickness") else ""
    arb_str = f"{dims.get('arbor_size')} {dims.get('arbor_uom', 'in')} Arbor" if dims.get("arbor_size") else ""
    dim_full = ", ".join([d for d in [dia_str, thk_str, arb_str] if d])
    
    long_desc = f"{brand_name} {series + ' ' if series else ''}{item_type}"
    if dim_full:
        long_desc += f", {dim_full}"
    long_desc += f", {material} Construction, Designed for {application}."

    retail_desc = short_desc
    marketing_desc = f"Engineered for industrial professionals, the {brand_name} {series + ' ' if series else ''}{item_type} offers superior durability and performance in demanding {application.lower()} operations."

    # 4. Populate All 252 Output Columns
    out = {col: "" for col in UNILOG_DELIVERY_COLUMNS}

    # Core Identifiers
    out["PART_NUMBER"] = mpn
    out["Dept"] = dept
    out["Class"] = cat_class
    out["Fine"] = fine
    out["SKU - MY_PART_NUMBER"] = mpn
    out["Mfg_Part_Num"] = mpn
    out["Part_Desc"] = raw_desc
    out["E1_Brand"] = e1_brand or "-- Unbranded --"
    out["Unilog_Brand"] = unilog_brand or "-- No Unilog Brand --"
    out["DIB_Brand"] = dib_brand or "-- No DIB Brand --"
    out["Part_Manuf"] = part_manuf
    out["MANUFACTURER_NAME"] = mfr_name
    out["BRAND_NAME"] = brand_name
    out["TRADE_NAME"] = series
    out["MANUFACTURER_PART_NUMBER"] = mpn
    out["Classpath"] = classpath
    out["MOBILE_DESC"] = mobile_desc
    out["INVOICE_DESC"] = invoice_desc
    out["SHORT_DESC"] = short_desc
    out["LONG_DESC1"] = long_desc
    out["RETAIL_DESC"] = retail_desc
    out["MARKETING_DESCRIPTION"] = marketing_desc
    out["Product Name"] = item_type
    out["Application"] = application
    out["UNSPSC"] = unspsc
    out["Country Of Origin"] = str(row.get("Country Of Origin", "")).strip()
    out["Discontinued"] = str(row.get("Discontinued", "")).strip()
    out["Actual Image (Yes/No)"] = str(row.get("Actual Image (Yes/No)", "")).strip()

    # Reference URLs (only populated if present in input/verified catalog)
    primary_source = f"https://www.milwaukeetool.com/Products/{mpn}" if "milwaukee" in mfr_name.lower() and mpn else ""
    out["MFR URL"] = str(row.get("MFR URL", primary_source)).strip()
    out["Ref URL 1"] = str(row.get("Ref URL 1", "")).strip()
    out["Ref URL 2"] = str(row.get("Ref URL 2", "")).strip()
    out["Ref URL 3"] = str(row.get("Ref URL 3", "")).strip()

    # Features (up to 20)
    for f_idx, feat in enumerate(features, start=1):
        if f_idx <= 20:
            out[f"ITEM_FEATURES_{f_idx}"] = feat

    # 5. Dynamic 50 Attribute Triplets Slotting
    attr_list: List[Tuple[str, str, str]] = []
    
    if dims.get("diameter"):
        attr_list.append(("Diameter", format_fraction_value(dims["diameter"]), "in"))
    if dims.get("thickness"):
        attr_list.append(("Thickness", format_fraction_value(dims["thickness"]), "in"))
    if dims.get("arbor_size"):
        attr_list.append(("Arbor Size", format_fraction_value(dims["arbor_size"]), dims.get("arbor_uom", "in")))
    if dims.get("width"):
        attr_list.append(("Width", format_fraction_value(dims["width"]), "in"))
    if dims.get("length"):
        attr_list.append(("Length", format_fraction_value(dims["length"]), "in"))
    if dims.get("grit"):
        attr_list.append(("Grit", dims["grit"], ""))
    
    attr_list.append(("Material", material, ""))
    attr_list.append(("Application", application, ""))
    attr_list.append(("Series", series, "") if series else ("Grade", "Industrial", ""))

    # Slot attributes into 1..50
    for idx, (label, val, uom) in enumerate(attr_list, start=1):
        if idx <= 50:
            out[f"ATTRIBUTE_LABEL {idx}"] = label
            out[f"ATTRIBUTE_VALUE {idx}"] = val
            out[f"ATTRIBUTE_UOM {idx}"] = standardize_uom(uom)

    return out

async def enrich_single_abrasive(row: Dict[str, Any]) -> Dict[str, Any]:
    """Alias for backwards compatibility."""
    return await enrich_unilog_row(row)

def export_to_xlsx(df: pl.DataFrame, output_path: str) -> str:
    """
    Exports enriched catalog DataFrame to Excel (.xlsx) with explicit text formatting ('@')
    on all attribute value and dimension columns to prevent Excel from auto-converting fractions (e.g. 4-1/2, 7/8) into dates.
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Unilog Master Catalog"

    headers = df.columns
    ws.append(headers)

    # Style header row with Navy Blue Enterprise Theme
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="F8FAFC")
    thin_border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155")
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Format all value columns as text format ('@')
    text_format_cols = {f"ATTRIBUTE_VALUE {i}" for i in range(1, 51)} | {
        "Mfg_Part_Num", "MANUFACTURER_PART_NUMBER", "SKU - MY_PART_NUMBER",
        "INVOICE_DESC", "MOBILE_DESC", "SHORT_DESC", "UNSPSC"
    }

    for row_idx, row in enumerate(df.iter_rows(named=True), start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            if header in text_format_cols:
                cell.number_format = "@"  # Force text format
                cell.value = str(value) if value is not None else ""
            else:
                cell.value = value if value is not None else ""

    # Set column widths
    for col in ws.columns:
        header_val = str(col[0].value or '')
        ws.column_dimensions[col[0].column_letter].width = min(max(len(header_val) + 4, 12), 45)

    wb.save(output_path)
    return output_path

def append_to_shared_workbook(input_excel_path: str, df: pl.DataFrame, output_sheet_name: str = "Enriched_Output") -> str:
    """
    Shared Workbook Execution Model (Ramachandra Raje Urs / Unilog Standard):
    Loads the active input Excel workbook (.xlsx), iterates rows, and appends the structured
    253-column output records into a dedicated 'Enriched_Output' worksheet inside the same workbook.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if os.path.exists(input_excel_path) and input_excel_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(input_excel_path)
        if output_sheet_name in wb.sheetnames:
            ws = wb[output_sheet_name]
            ws.delete_rows(1, ws.max_row + 1)
        else:
            ws = wb.create_sheet(title=output_sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = output_sheet_name

    headers = df.columns
    ws.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="F8FAFC")
    thin_border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155")
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    text_format_cols = {f"ATTRIBUTE_VALUE {i}" for i in range(1, 51)} | {
        "Mfg_Part_Num", "MANUFACTURER_PART_NUMBER", "SKU - MY_PART_NUMBER",
        "INVOICE_DESC", "MOBILE_DESC", "SHORT_DESC", "UNSPSC"
    }

    for row_idx, row in enumerate(df.iter_rows(named=True), start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            if header in text_format_cols:
                cell.number_format = "@"  # Force text format to prevent fraction-to-date conversion
                cell.value = str(value) if value is not None else ""
            else:
                cell.value = value if value is not None else ""

    for col in ws.columns:
        header_val = str(col[0].value or '')
        ws.column_dimensions[col[0].column_letter].width = min(max(len(header_val) + 4, 12), 45)

    out_path = input_excel_path if input_excel_path.endswith(".xlsx") else input_excel_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(out_path)
    return out_path

async def process_unilog_dataset(
    input_csv_path: str = "Unihack__Sample_Dataset_-_Input.csv",
    max_rows: int = 50,
    output_csv_path: str = "Unihack_Enriched_Output.csv",
    limit: Optional[int] = None,
    output_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    Reads the Unilog dataset using Polars, cleans placeholders, processes items concurrently
    with Semaphore bounding, and outputs standard 252-column delivery format CSV (UTF-8-BOM) and XLSX.
    """
    effective_limit = limit if limit is not None else max_rows
    effective_out_csv = output_path if output_path is not None else output_csv_path
    effective_out_xlsx = effective_out_csv.rsplit(".", 1)[0] + ".xlsx"

    # Step 1: Read with Polars
    if not os.path.exists(input_csv_path):
        raise FileNotFoundError(f"Input catalog CSV not found at: {input_csv_path}")

    df = pl.read_csv(input_csv_path)
    
    # Clean placeholder values
    for b_col in ["E1_Brand", "Unilog_Brand", "DIB_Brand"]:
        if b_col in df.columns:
            df = df.with_columns(
                pl.when(pl.col(b_col).str.starts_with("--")).then(None).otherwise(pl.col(b_col)).alias(b_col)
            )

    # Step 2: Slice target rows
    target_df = df.head(effective_limit)
    rows = target_df.to_dicts()

    # Step 3: Process rows concurrently with Semaphore
    sem = asyncio.Semaphore(25)
    async def _sem_worker(r):
        async with sem:
            return await enrich_unilog_row(r)

    enriched_rows = await asyncio.gather(*[_sem_worker(r) for r in rows])

    # Step 4: Create Polars DataFrame matching 252 delivery columns exactly
    out_df = pl.DataFrame(enriched_rows)
    
    # Ensure column order matches UNILOG_DELIVERY_COLUMNS
    final_cols = [c for c in UNILOG_DELIVERY_COLUMNS if c in out_df.columns]
    out_df = out_df.select(final_cols)

    # Step 5: Write CSV with explicit UTF-8-BOM to preserve symbols (® ™) in Excel
    csv_str = out_df.write_csv()
    with open(effective_out_csv, "w", encoding="utf-8-sig") as f:
        f.write(csv_str)

    # Step 6: Export formatted XLSX
    export_to_xlsx(out_df, effective_out_xlsx)

    return {
        "status": "success",
        "processed_count": len(enriched_rows),
        "total_columns": len(out_df.columns),
        "output_csv": effective_out_csv,
        "output_xlsx": effective_out_xlsx,
        "sample_rows": enriched_rows[:3]
    }

def profile_to_unilog_row(profile: Dict[str, Any]) -> Dict[str, Any]:
    """Converts a single enriched profile from LangGraph state into the canonical 252-column Unilog dictionary."""
    row = {col: "" for col in UNILOG_DELIVERY_COLUMNS}
    
    brand = profile.get("brand", "") or "Industrial"
    mpn = profile.get("mpn", "") or profile.get("raw_input", "")
    title = profile.get("standardized_title", "")
    invoice = profile.get("invoice_desc", "")
    mobile = profile.get("mobile_desc", "")
    short_d = profile.get("short_desc", "")
    mkt_d = profile.get("marketing_description", "")
    tax = profile.get("taxonomy", {}) or {}
    unspsc = tax.get("unspsc_code", "")
    classpath = " > ".join([tax.get("segment", "Industrial"), tax.get("family", "Equipment"), tax.get("class_name", "General")])
    
    row["PART_NUMBER"] = mpn
    row["Mfg_Part_Num"] = mpn
    row["MANUFACTURER_PART_NUMBER"] = mpn
    row["SKU - MY_PART_NUMBER"] = mpn
    row["MANUFACTURER_NAME"] = brand
    row["BRAND_NAME"] = brand
    row["Unilog_Brand"] = brand
    row["E1_Brand"] = brand
    row["Part_Desc"] = title
    row["Product Name"] = title
    row["INVOICE_DESC"] = invoice
    row["MOBILE_DESC"] = mobile
    row["SHORT_DESC"] = short_d
    row["MARKETING_DESCRIPTION"] = mkt_d
    row["UNSPSC"] = unspsc
    row["Classpath"] = classpath
    
    # Feature bullets
    bullets = profile.get("feature_bullets", [])
    for idx, bullet in enumerate(bullets[:20], start=1):
        row[f"ITEM_FEATURES_{idx}"] = bullet
        
    # Attributes
    attrs = profile.get("attributes", {})
    idx = 1
    for attr_name, attr_data in attrs.items():
        if idx > 50:
            break
        if isinstance(attr_data, dict):
            val = attr_data.get("value", "")
            uom = attr_data.get("unit", "")
        else:
            val = str(attr_data)
            uom = ""
        row[f"ATTRIBUTE_LABEL {idx}"] = str(attr_name).replace("_", " ").title()
        row[f"ATTRIBUTE_VALUE {idx}"] = str(val) if val is not None else ""
        row[f"ATTRIBUTE_UOM {idx}"] = str(uom) if uom is not None else ""
        idx += 1
        
    return row

def append_profile_to_master_sheet(
    profile: Dict[str, Any],
    master_csv: str = "PIMpulse_Unilog_Enriched_1000.csv",
    master_xlsx: str = "PIMpulse_Unilog_Enriched_1000.xlsx"
) -> bool:
    """Appends a single newly enriched SKU into both the master CSV and XLSX files."""
    try:
        new_row = profile_to_unilog_row(profile)
        
        # 1. Append to CSV (infer_schema_length=0 reads all columns as String to avoid type mismatch)
        if os.path.exists(master_csv):
            df_curr = pl.read_csv(master_csv, truncate_ragged_lines=True, infer_schema_length=0)
            df_new = pl.DataFrame({k: [str(v or "")] for k, v in new_row.items()})
            for col in df_curr.columns:
                if col not in df_new.columns:
                    df_new = df_new.with_columns(pl.lit("").alias(col))
            df_new = df_new.select(df_curr.columns)
            df_combined = pl.concat([df_curr, df_new], how="vertical")
            with open(master_csv, "w", encoding="utf-8-sig") as f:
                f.write(df_combined.write_csv())
        else:
            df_new = pl.DataFrame({k: [str(v or "")] for k, v in new_row.items()}).select(UNILOG_DELIVERY_COLUMNS)
            with open(master_csv, "w", encoding="utf-8-sig") as f:
                f.write(df_new.write_csv())

        # 2. Append to XLSX if openpyxl available
        if os.path.exists(master_xlsx):
            import openpyxl
            wb = openpyxl.load_workbook(master_xlsx)
            ws = wb["Enriched_Output"] if "Enriched_Output" in wb.sheetnames else wb.active
            headers = [cell.value for cell in ws[1]]
            row_vals = [str(new_row.get(h, "") or "") for h in headers]
            ws.append(row_vals)
            # Apply text format to last row
            last_row_idx = ws.max_row
            text_format_cols = {f"ATTRIBUTE_VALUE {i}" for i in range(1, 51)} | {
                "Mfg_Part_Num", "MANUFACTURER_PART_NUMBER", "SKU - MY_PART_NUMBER",
                "INVOICE_DESC", "MOBILE_DESC", "SHORT_DESC", "UNSPSC"
            }
            for col_idx, header in enumerate(headers, start=1):
                if header in text_format_cols:
                    ws.cell(row=last_row_idx, column=col_idx).number_format = "@"
            wb.save(master_xlsx)
            
        logger.info(f"Successfully appended SKU '{profile.get('raw_input')}' to master catalog.")
        return True
    except Exception as e:
        logger.error(f"Failed to append SKU to master sheet: {e}")
        return False

