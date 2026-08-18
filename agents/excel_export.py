import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from typing import List, Dict, Any, Optional

def export_to_excel(results: List[Dict[str, Any]], input_file: Optional[str] = None, output_path: str = "enriched_catalog.xlsx") -> str:
    """
    Takes pipeline results and writes enriched industrial product profiles to Excel.
    If input_file provided, adds output as Sheet 2 ('PIMpulse_Enriched') in the same workbook.
    Otherwise creates a new styled workbook.
    """
    if input_file and Path(input_file).exists():
        wb = openpyxl.load_workbook(input_file)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet if empty
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])

    # Create or replace enriched output sheet
    ws = wb.create_sheet("PIMpulse_Enriched") if "PIMpulse_Enriched" not in wb.sheetnames else wb["PIMpulse_Enriched"]
    ws.delete_rows(1, ws.max_row + 10)

    # Column Headers
    headers = [
        "Raw Input", "Standardized Title", "UNSPSC Code",
        "Category Path", "Confidence %", "Decision",
        "Material", "Thread Size", "Length", "Voltage",
        "Current Rating", "Poles", "Bore Diameter", "Power Rating",
        "Marketing Description", "Source Provenance URLs", "Ambiguity Flag"
    ]

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, color="00FF88", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data Rows
    row_font = Font(name="Segoe UI", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for row_idx, result in enumerate(results, 2):
        attrs = result.get("attributes", {})
        taxonomy = result.get("taxonomy", {})
        confidence = result.get("confidence", {})
        provenance = result.get("provenance", {})

        def get_attr_val(key: str) -> str:
            val = attrs.get(key, {})
            if isinstance(val, dict):
                v = str(val.get("value", "") or "")
                u = val.get("unit")
                return f"{v} {u}".strip() if u and u not in v else v
            return str(val or "")

        raw_inp = result.get("raw_input", "")
        title = result.get("standardized_title", "")
        code = str(taxonomy.get("code", "00000000"))
        path_str = " > ".join(taxonomy.get("path", [])) if taxonomy.get("path") else taxonomy.get("class_name", "")
        conf_pct = confidence.get("confidence_pct", 0)
        decision = str(result.get("evaluator_decision", "ACCEPT")).upper()
        
        # Sources
        sources = []
        if isinstance(provenance, dict):
            sources = list(set([str(v) for v in provenance.values() if str(v).startswith("http")]))
        source_str = " | ".join(sources)[:500]

        row_data = [
            raw_inp,
            title,
            code,
            path_str,
            f"{conf_pct}%",
            decision,
            get_attr_val("material"),
            get_attr_val("thread_size"),
            get_attr_val("length"),
            get_attr_val("voltage") or get_attr_val("supply_voltage") or get_attr_val("coil_voltage"),
            get_attr_val("current_rating") or get_attr_val("rated_current"),
            get_attr_val("poles"),
            get_attr_val("bore_diameter"),
            get_attr_val("power_rating") or get_attr_val("rated_power"),
            result.get("marketing_description", ""),
            source_str,
            "TRUE" if result.get("ambiguity_flag") else "FALSE"
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.border = thin_border
            if col_idx in (3, 5, 6, 17):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    ws.row_dimensions[1].height = 28
    wb.save(output_path)
    return output_path
