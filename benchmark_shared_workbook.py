"""
Benchmark script validating the Shared Workbook Execution Model (Ramachandra Raje Urs / Unilog Standard)
Loads an input .xlsx spreadsheet, processes rows, and appends the 253-column structured data
into a dedicated 'Enriched_Output' worksheet inside the same Excel workbook.
"""

import os
import polars as pl
import openpyxl
import asyncio
from agents.unilog_pipeline import enrich_unilog_row, append_to_shared_workbook, UNILOG_DELIVERY_COLUMNS

async def main():
    test_wb_path = "test_shared_workbook.xlsx"
    
    # 1. Create a dummy input workbook with a raw input sheet
    wb = openpyxl.Workbook()
    ws_input = wb.active
    ws_input.title = "Raw_Input_SKUs"
    ws_input.append(["Mfg_Part_Num", "Part_Desc", "Part_Manuf"])
    
    sample_inputs = [
        ("49-94-0107", "4-1/2 IN. X .045 IN. X 7/8 IN. METAL CUT-OFF WHEEL (4031)", "Milwaukee Accessory (4031)"),
        ("23-612-180", "5inch Abranet Grip Mesh Disc 180 Grit Alum Oxide", "Mirka Abrasives (MIRK)"),
        ("3RT2015-1BB41", "SIRIUS power contactor 3P 24VDC 4kW S00", "Siemens Industry"),
        ("6205-2RSH", "Deep groove ball bearing 25x52x15 mm rubber sealed", "SKF"),
        ("BLT-316-1/2-13", "1/2-13 x 2-1/2 Grade 316 Stainless Steel Heavy Hex Bolt", "McMaster-Carr")
    ]
    for row in sample_inputs:
        ws_input.append(list(row))
        
    wb.save(test_wb_path)
    print(f"1. Created shared input workbook '{test_wb_path}' with sheet '{ws_input.title}'.")
    
    # 2. Enrich the rows
    enriched = []
    for mpn, desc, mfr in sample_inputs:
        res = await enrich_unilog_row({"Mfg_Part_Num": mpn, "Part_Desc": desc, "Part_Manuf": mfr})
        enriched.append(res)
        
    df_out = pl.DataFrame(enriched)
    final_cols = [c for c in UNILOG_DELIVERY_COLUMNS if c in df_out.columns]
    df_out = df_out.select(final_cols)
    
    # 3. Append to shared workbook
    out_path = append_to_shared_workbook(test_wb_path, df_out, "Enriched_Output")
    print(f"2. Successfully appended 'Enriched_Output' sheet to shared workbook '{out_path}'.")
    
    # 4. Verify sheets inside the workbook
    wb_verify = openpyxl.load_workbook(test_wb_path)
    sheet_names = wb_verify.sheetnames
    print(f"3. Worksheets present in workbook: {sheet_names}")
    assert "Raw_Input_SKUs" in sheet_names
    assert "Enriched_Output" in sheet_names
    
    ws_out = wb_verify["Enriched_Output"]
    print(f"4. Enriched_Output rows: {ws_out.max_row}, cols: {ws_out.max_column}")
    assert ws_out.max_row == len(sample_inputs) + 1
    assert ws_out.max_column == len(final_cols)
    
    # Clean up test file
    if os.path.exists(test_wb_path):
        os.remove(test_wb_path)
    print("5. Shared Workbook Execution Model validation complete and verified!\n")

if __name__ == "__main__":
    asyncio.run(main())
